import streamlit as st  # Streamlit 라이브러리 불러오기 - 웹 인터페이스 구축
from openpyxl import Workbook, load_workbook  # 엑셀 파일 처리를 위한 라이브러리
from openpyxl.styles import Font, Alignment  # 엑셀 셀 서식 지정용 스타일 클래스
from datetime import datetime  # 날짜 및 시간 처리를 위한 라이브러리
import os  # 파일 및 디렉토리 조작을 위한 라이브러리
import pandas as pd  # 데이터 처리를 위한 라이브러리
import gspread  # Google Sheets API 연동
from google.oauth2.service_account import Credentials  # Google API 인증
from io import BytesIO  # 메모리 내 파일 처리
import logging  # 로깅을 위한 라이브러리
import time  # 시간 처리를 위한 라이브러리

# Google Sheets API 설정
def setup_google_sheets():
    try:
        scope = [
            'https://spreadsheets.google.com/feeds',
            'https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/drive'
        ]
        
        # Streamlit 로컬 개발 환경인지 클라우드 환경인지 확인
        if os.path.exists('.streamlit/secrets.toml'):
            # 로컬 개발 환경일 경우
            credentials = Credentials.from_service_account_info(
                st.secrets["gcp_service_account"],
                scopes=scope
            )
        else:
            try:
                # Streamlit Cloud 환경일 경우
                credentials = Credentials.from_service_account_info(
                    st.secrets["gcp_service_account"],
                    scopes=scope
                )
            except Exception as e:
                st.error(f"Google API 인증 정보를 찾을 수 없습니다: {e}")
                st.info("관리자에게 문의하세요. Streamlit Secrets에 서비스 계정 정보가 필요합니다.")
                return None
        
        client = gspread.authorize(credentials)
        return client
    except Exception as e:
        st.error(f"Google Sheets API 설정 중 오류가 발생했습니다: {e}")
        return None

# 스프레드시트 열기 또는 생성
def get_or_create_spreadsheet(client, sheet_name):
    try:
        # 스프레드시트 열기 시도
        spreadsheet = client.open(sheet_name)
        st.info(f"기존 스프레드시트를 열었습니다: {sheet_name}")
        
        # 기존 스프레드시트에도 권한 부여 시도
        try:
            # 현재 사용자 이메일 주소
            user_email = 'qhv147@gmail.com'
            
            # 이미 접근 권한이 있는지 확인 (API 호출을 줄이기 위함)
            try:
                # 현재 권한 목록 가져오기
                permissions = spreadsheet.list_permissions()
                existing_emails = [p.get('emailAddress', '') for p in permissions]
                
                # 이미 권한이 있으면 건너뛰기
                if user_email in existing_emails:
                    st.info("이미 스프레드시트에 대한 접근 권한이 있습니다.")
                else:
                    # 권한 부여 시도
                    spreadsheet.share(user_email, perm_type='user', role='writer')
                    st.success("기존 스프레드시트에 접근 권한이 부여되었습니다.")
            except:
                # 권한 목록 조회 실패 시 그냥 공유 시도
                spreadsheet.share(user_email, perm_type='user', role='writer')
                st.success("기존 스프레드시트에 접근 권한이 부여되었습니다.")
                
        except Exception as e:
            st.warning(f"기존 스프레드시트 공유 중 오류가 발생했습니다: {str(e)[:100]}... 스프레드시트 소유자에게 권한을 요청하세요.")
            # 스프레드시트 URL과 함께 자세한 안내 제공
            st.info(f"이 스프레드시트({sheet_name})에 접근하려면 소유자에게 '{user_email}' 계정에 대한 권한을 요청하세요.")
            
    except gspread.exceptions.SpreadsheetNotFound:
        # 스프레드시트가 없으면 새로 생성
        spreadsheet = client.create(sheet_name)
        st.success(f"새 스프레드시트를 생성했습니다: {sheet_name}")
        
        # 새로 생성된 스프레드시트에 공유 시도
        try:
            # 기본 권한 설정 - 자신에게 편집 권한 부여
            spreadsheet.share('qhv147@gmail.com', perm_type='user', role='writer')
            st.success("새 스프레드시트에 접근 권한이 부여되었습니다.")
        except Exception as e:
            st.warning(f"스프레드시트 공유 중 오류가 발생했습니다: {str(e)[:100]}... 나중에 수동으로 공유해주세요.")
    
    return spreadsheet

# 워크시트 가져오기 또는 생성
def get_or_create_worksheet(spreadsheet, worksheet_name):
    try:
        # 워크시트 열기 시도
        worksheet = spreadsheet.worksheet(worksheet_name)
    except gspread.exceptions.WorksheetNotFound:
        # 워크시트가 없으면 새로 생성
        worksheet = spreadsheet.add_worksheet(title=worksheet_name, rows=1000, cols=20)
        
        # 헤더 추가
        headers = [
            "NO", "월", "구분", "작업유형", "TASK", "요청일", "작업일",
            "요청자", "IT", "CNS", "개발자", "내용", "결과"
        ]
        worksheet.append_row(headers)
        
        # 열 너비 설정 (Google Sheets API에서는 직접 지원하지 않음)
    
    return worksheet

# 현업문의 워크시트 가져오기 또는 생성
def get_or_create_inquiry_worksheet(spreadsheet, worksheet_name):
    try:
        # 워크시트 열기 시도
        worksheet = spreadsheet.worksheet(worksheet_name)
    except gspread.exceptions.WorksheetNotFound:
        # 워크시트가 없으면 새로 생성
        worksheet = spreadsheet.add_worksheet(title=worksheet_name, rows=1000, cols=20)
        
        # 헤더 추가
        headers = [
            "NO", "월", "문의방법", "문의유형", "요청부서", "문의사항", "요청일", "답변일",
            "요청자", "IT", "CNS", "개발자"
        ]
        worksheet.append_row(headers)
        
        # 열 너비 설정 (Google Sheets API에서는 직접 지원하지 않음)
    
    return worksheet

# 데이터 정렬 함수 (요청일 기준)
def sort_worksheet_by_date(worksheet, date_col_idx=5):
    """
    날짜 기준으로 워크시트 데이터를 정렬합니다.
    date_col_idx: 정렬 기준이 될 날짜 열의 인덱스 (기본값: 5, 요청일 열)
    """
    try:
        # 모든 데이터 가져오기
        data = worksheet.get_all_values()
        
        # 헤더 제외하고 데이터만 가져오기
        headers = data[0]
        data_rows = data[1:]
        
        # 데이터 없으면 바로 반환
        if not data_rows:
            return
            
        # 이미 정렬되어 있는지 확인
        # date_col_idx = 5  # 요청일 열 인덱스 (6번째 열, 0부터 시작)
        
        # 날짜 형식 변환 함수
        def parse_date(date_str):
            try:
                # '2023-12-31' 형식의 날짜 처리
                return datetime.strptime(date_str, '%Y-%m-%d')
            except ValueError:
                try:
                    # '23-12-31' 등의 2자리 연도 형식 처리
                    parsed_date = datetime.strptime(date_str, '%y-%m-%d')
                    # 2000년 이전인지 확인 및 조정
                    current_year = datetime.now().year
                    century = (current_year // 100) * 100
                    if parsed_date.year > (current_year % 100):
                        # 과거 날짜로 가정
                        parsed_date = parsed_date.replace(year=parsed_date.year + century - 100)
                    else:
                        # 현재 세기로 가정
                        parsed_date = parsed_date.replace(year=parsed_date.year + century)
                    return parsed_date
                except ValueError:
                    # 다른 형식이거나 유효하지 않은 날짜는 매우 오래된 날짜로 처리
                    return datetime(1900, 1, 1)
        
        # 정렬 필요 여부 확인
        is_sorted = True
        for i in range(1, len(data_rows)):
            prev_date = parse_date(data_rows[i-1][date_col_idx])
            curr_date = parse_date(data_rows[i][date_col_idx])
            if prev_date > curr_date:
                is_sorted = False
                break
                
        if is_sorted:
            return  # 이미 정렬되어 있음
        
        # 날짜 기준으로 정렬 (오래된 날짜가 위로)
        sorted_data = sorted(data_rows, key=lambda x: parse_date(x[date_col_idx]))
        
        # 배치 업데이트를 위한 준비
        batch_size = 100  # 한 번에 업데이트할 최대 행 수
        total_batches = (len(sorted_data) + batch_size - 1) // batch_size
        
        # 헤더는 그대로 두고 정렬된 데이터만 업데이트
        for i in range(0, len(sorted_data), batch_size):
            batch = sorted_data[i:i+batch_size]
            start_row = i + 2  # 헤더(1) + 데이터 시작 인덱스(i+1)
            
            # 배치 단위로 업데이트 - 인자 순서 수정
            worksheet.update(values=batch, range_name=f'A{start_row}')
            
            # API 할당량 제한을 고려한 딜레이
            if i + batch_size < len(sorted_data):
                time.sleep(2)  # 2초 대기
                
        return True
        
    except Exception as e:
        logging.error(f"워크시트 정렬 중 오류: {str(e)}")
        raise Exception(f"데이터 정렬 중 오류가 발생했습니다: {str(e)}")

# 캐싱을 위한 데코레이터 추가
@st.cache_data(ttl=300)  # 5분 동안 결과 캐싱
def get_worksheet_data(_worksheet):
    """
    워크시트의 데이터를 가져와 캐싱합니다.
    이 함수는 동일한 워크시트에 대해 짧은 시간 내에 반복 호출될 경우 
    API 호출 없이 캐시된 데이터를 반환합니다.
    """
    # 디버깅을 위한 로그 추가
    st.session_state['last_data_fetch'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return _worksheet.get_all_values()

# 업로드 후 캐시를 명시적으로 갱신하는 함수
def refresh_worksheet_data():
    """
    데이터 업로드/추가 후 워크시트 캐시를 갱신하는 함수
    """
    # 캐시 무효화
    get_worksheet_data.clear()
    st.session_state['cache_refreshed'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # 페이지 자동 새로고침을 위한 플래그
    st.session_state['data_updated'] = True
    
# 성능 최적화를 위한 함수 추가
@st.cache_data(ttl=600)  # 10분 동안 캐싱
def get_spreadsheet_info(gs_client, sheet_name):
    """스프레드시트 정보를 가져오고 캐싱합니다."""
    try:
        sheet = gs_client.open(sheet_name)
        return {
            "url": sheet.url,
            "exists": True
        }
    except gspread.exceptions.SpreadsheetNotFound:
        return {
            "url": "#",
            "exists": False
        }

# 세션 상태 초기화 - 요청일과 작업일 동기화를 위한 설정
if 'req_date' not in st.session_state:
    st.session_state.req_date = datetime.today()
    st.session_state.prev_req_date = datetime.today()

if 'work_date' not in st.session_state:
    st.session_state.work_date = datetime.today()

# 현업문의 요청일과 답변일 동기화를 위한 설정
if 'inquiry_req_date' not in st.session_state:
    st.session_state.inquiry_req_date = datetime.today()
    st.session_state.prev_inquiry_req_date = datetime.today()

if 'inquiry_resp_date' not in st.session_state:
    st.session_state.inquiry_resp_date = datetime.today()

# 데이터 캐시 관련 세션 상태 초기화
if 'last_data_fetch' not in st.session_state:
    st.session_state.last_data_fetch = None

if 'cache_refreshed' not in st.session_state:
    st.session_state.cache_refreshed = None

if 'data_updated' not in st.session_state:
    st.session_state.data_updated = False

# 페이지 자동 새로고침을 위한 처리
if st.session_state.get('data_updated', False):
    # 플래그 초기화
    st.session_state.data_updated = False
    # 캐시 초기화
    get_worksheet_data.clear()

# 최적화된 요청일 변경 콜백 함수
def update_work_date():
    """
    요청일이 변경될 때 작업일을 업데이트합니다.
    불필요한 재계산이나 API 호출을 하지 않습니다.
    """
    # 이전 값과 동일하면 아무 작업도 수행하지 않음
    if 'prev_req_date' in st.session_state and st.session_state.prev_req_date == st.session_state.req_date:
        return

    # 현재 값을 저장
    st.session_state.prev_req_date = st.session_state.req_date
    st.session_state.work_date = st.session_state.req_date
    # 상태 변경만 수행하고 추가적인 처리는 하지 않음

# 현업문의 요청일 변경 콜백 함수도 최적화
def update_inquiry_resp_date():
    """요청일이 변경될 때 답변일을 업데이트합니다."""
    # 이전 값과 동일하면 아무 작업도 수행하지 않음
    if 'prev_inquiry_req_date' in st.session_state and st.session_state.prev_inquiry_req_date == st.session_state.inquiry_req_date:
        return

    # 현재 값을 저장
    st.session_state.prev_inquiry_req_date = st.session_state.inquiry_req_date
    st.session_state.inquiry_resp_date = st.session_state.inquiry_req_date

# Streamlit UI - 웹 애플리케이션 제목 설정
st.title("🛠 SM Activity 기록 프로그램")

# Google Sheets API 클라이언트 초기화
gs_client = setup_google_sheets()
if not gs_client:
    st.error("Google Sheets API에 연결할 수 없습니다.")
    st.stop()

# 파일 선택 옵션 - 사용자가 선택할 수 있는 스프레드시트 옵션 정의
sheet_options = {
    "SM Activity - 대시보드": "SM Activity Dashboard",
    "SM Activity - Plan": "SM Activity Plan"
}

# 사용자가 작업할 스프레드시트 선택을 위한 드롭다운 생성
selected_sheet_name = st.selectbox(
    "작성할 문서 선택", 
    options=list(sheet_options.keys())
)

# 선택된 스프레드시트 이름 설정
google_sheet_name = sheet_options[selected_sheet_name]
worksheet_name = "SM Activity"  # 모든 시트에 동일한 워크시트 이름 사용
inquiry_worksheet_name = "현업문의"  # 현업문의 워크시트 이름

# 선택한 스프레드시트 열기 또는 생성
spreadsheet = get_or_create_spreadsheet(gs_client, google_sheet_name)
if not spreadsheet:
    st.error("스프레드시트에 접근할 수 없습니다.")
    st.stop()

# 워크시트 열기 또는 생성
worksheet = get_or_create_worksheet(spreadsheet, worksheet_name)
if not worksheet:
    st.error("워크시트에 접근할 수 없습니다.")
    st.stop()

# 현업문의 워크시트 열기 또는 생성
inquiry_worksheet = get_or_create_inquiry_worksheet(spreadsheet, inquiry_worksheet_name)
if not inquiry_worksheet:
    st.error("현업문의 워크시트에 접근할 수 없습니다.")
    st.stop()

# 스프레드시트 링크 항상 표시
st.markdown(f"### 📊 Google 스프레드시트")
st.markdown(f"""
<div style='background-color: #f0f2f6; padding: 15px; border-radius: 5px; margin-bottom: 15px;'>
    <p><strong>현재 선택된 스프레드시트:</strong> <a href='{spreadsheet.url}' target='_blank'>{google_sheet_name}</a></p>
    
    <details>
        <summary><strong>모든 스프레드시트 링크</strong></summary>
        <ul>
""", unsafe_allow_html=True)

# 모든 스프레드시트 링크 표시
for sheet_label, sheet_name in sheet_options.items():
    # 현재 선택된 시트인지 확인
    is_current = sheet_name == google_sheet_name
    try:
        # 최적화된 함수를 사용하여 스프레드시트 정보 가져오기
        sheet_info = get_spreadsheet_info(gs_client, sheet_name)
        
        if sheet_info["exists"]:
            if is_current:
                st.markdown(f"<li><strong>{sheet_label}</strong>: <a href='{sheet_info['url']}' target='_blank'>{sheet_name}</a> (현재 선택됨)</li>", unsafe_allow_html=True)
            else:
                st.markdown(f"<li><strong>{sheet_label}</strong>: <a href='{sheet_info['url']}' target='_blank'>{sheet_name}</a></li>", unsafe_allow_html=True)
        else:
            st.markdown(f"<li><strong>{sheet_label}</strong>: {sheet_name} (아직 생성되지 않음)</li>", unsafe_allow_html=True)
    except Exception as e:
        st.markdown(f"<li><strong>{sheet_label}</strong>: {sheet_name} (링크 확인 중 오류 발생)</li>", unsafe_allow_html=True)

st.markdown("""
        </ul>
    </details>
    
    <p><small>만약 접근 권한이 없다면 다시 앱을 로드하거나, 스프레드시트 소유자에게 권한을 요청하세요.</small></p>
</div>
""", unsafe_allow_html=True)

# 디버깅 정보를 사이드바에 추가
with st.sidebar:
    st.subheader("🧩 시스템 정보")
    with st.expander("캐시 및 데이터 상태"):
        st.write(f"마지막 데이터 조회: {st.session_state.get('last_data_fetch', '없음')}")
        st.write(f"마지막 캐시 갱신: {st.session_state.get('cache_refreshed', '없음')}")
        if st.button("캐시 수동 갱신"):
            get_worksheet_data.clear()
            st.session_state.cache_refreshed = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            st.success("✅ 캐시가 갱신되었습니다.")
            st.rerun()

# 통합 다운로드 버튼 섹션 추가
st.subheader("📥 데이터 다운로드")
with st.expander("모든 데이터 다운로드"):
    # 모든 시트의 데이터를 하나의 엑셀 파일로 다운로드
    st.markdown("현재 선택된 스프레드시트의 모든 데이터를 하나의 엑셀 파일로 다운로드할 수 있습니다.")
    
    # 통합 다운로드 버튼
    if st.button("전체 데이터 엑셀 파일 다운로드"):
        try:
            with st.spinner("엑셀 파일 생성 중..."):
                # 엑셀 파일 생성
                excel_buffer = BytesIO()
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    # SM Activity 데이터 가져오기
                    activity_data = get_worksheet_data(worksheet)
                    if len(activity_data) > 0:
                        activity_df = pd.DataFrame(activity_data[1:], columns=activity_data[0])
                        activity_df.to_excel(writer, index=False, sheet_name=worksheet_name)
                        
                        # 엑셀 서식 설정 - SM Activity 시트
                        workbook = writer.book
                        worksheet_excel = writer.sheets[worksheet_name]
                        
                        # 헤더 스타일 설정
                        for col_num, value in enumerate(activity_df.columns.values, 1):
                            cell = worksheet_excel.cell(row=1, column=col_num)
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # 열 너비 설정
                        worksheet_excel.column_dimensions['E'].width = 30  # TASK 컬럼
                        worksheet_excel.column_dimensions['F'].width = 15  # 요청일 컬럼
                        worksheet_excel.column_dimensions['G'].width = 15  # 작업일 컬럼
                        worksheet_excel.column_dimensions['L'].width = 40  # 내용 컬럼
                    
                    # 현업문의 데이터 가져오기
                    inquiry_data = get_worksheet_data(inquiry_worksheet)
                    if len(inquiry_data) > 0:
                        inquiry_df = pd.DataFrame(inquiry_data[1:], columns=inquiry_data[0])
                        inquiry_df.to_excel(writer, index=False, sheet_name=inquiry_worksheet_name)
                        
                        # 엑셀 서식 설정 - 현업문의 시트
                        workbook = writer.book
                        inquiry_worksheet_excel = writer.sheets[inquiry_worksheet_name]
                        
                        # 헤더 스타일 설정
                        for col_num, value in enumerate(inquiry_df.columns.values, 1):
                            cell = inquiry_worksheet_excel.cell(row=1, column=col_num)
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # 열 너비 설정
                        inquiry_worksheet_excel.column_dimensions['E'].width = 20  # 요청부서 컬럼
                        inquiry_worksheet_excel.column_dimensions['F'].width = 40  # 문의사항 컬럼
                        inquiry_worksheet_excel.column_dimensions['G'].width = 15  # 요청일 컬럼
                        inquiry_worksheet_excel.column_dimensions['H'].width = 15  # 답변일 컬럼
                
                excel_buffer.seek(0)
                
                # 다운로드 버튼 생성
                download_filename = f"{google_sheet_name}_통합데이터.xlsx"
                st.download_button(
                    label="📥 통합 엑셀 파일 다운로드",
                    data=excel_buffer,
                    file_name=download_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                st.success(f"✅ 통합 엑셀 파일이 생성되었습니다. 위 버튼을 클릭하여 다운로드하세요.")
                st.info(f"📊 파일 정보: SM Activity 및 현업문의 데이터가 각각 별도의 시트에 포함되어 있습니다.")
        except Exception as e:
            st.error(f"엑셀 파일 생성 중 오류가 발생했습니다: {str(e)[:200]}...")

# 탭 인터페이스 생성
tab1, tab2 = st.tabs(["SM Activity", "현업문의"])

# 세션 상태를 사용하여 현재 선택된 탭 추적
if 'current_tab' not in st.session_state:
    st.session_state.current_tab = "SM Activity"

# 탭이 변경되면 호출되는 콜백 함수
def update_current_tab(tab_name):
    st.session_state.current_tab = tab_name

with tab1:
    # SM Activity 탭을 선택했음을 세션 상태에 저장
    update_current_tab("SM Activity")
    
    # SM Activity 탭 내용
    # 엑셀 파일 업로드 섹션 추가
    st.subheader("📤 엑셀 파일 업로드")
    with st.expander("엑셀 파일을 업로드하여 데이터 일괄 추가"):
        # 샘플 템플릿 다운로드 기능 추가
        st.markdown("#### 샘플 템플릿 다운로드")
        sample_df = pd.DataFrame({
            '구분': ['정기', '비정기'],
            '작업유형': ['조간점검', '인프라 작업'],
            'TASK': ['데일리 점검', '서버 업그레이드'],
            '요청일': [datetime.today().strftime("%Y-%m-%d"), (datetime.today() - pd.Timedelta(days=1)).strftime("%Y-%m-%d")],
            '요청자': ['홍길동', '김철수'],
            'IT': ['한상욱', '한상욱'],
            'CNS': ['이정인', '이정인'],
            '개발자': ['위승빈', '위승빈'],
            '결과': ['완료', '진행 중']
        })
        
        # 샘플 템플릿을 엑셀로 변환
        sample_buffer = BytesIO()
        with pd.ExcelWriter(sample_buffer, engine='openpyxl') as writer:
            sample_df.to_excel(writer, index=False, sheet_name='SM Activity')
        sample_buffer.seek(0)
        
        # 샘플 템플릿 다운로드 버튼
        st.download_button(
            label="📝 샘플 템플릿 다운로드",
            data=sample_buffer,
            file_name="SM_Activity_Template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="업로드 양식에 맞는 샘플 엑셀 템플릿을 다운로드합니다."
        )
        
        st.markdown("---")
        st.markdown("#### 데이터 업로드")
        uploaded_file = st.file_uploader("SM Activity 양식의 엑셀 파일을 업로드하세요", type=["xlsx", "xls"], key="sm_activity_uploader")
        
        if uploaded_file is not None:
            try:
                # 엑셀 파일 읽기
                df = pd.read_excel(uploaded_file, sheet_name=0)
                
                # 데이터프레임 미리보기 
                st.write("업로드한 데이터 미리보기:")
                st.dataframe(df.head(5))
                
                # 필요한 열이 있는지 확인
                required_columns = ["구분", "작업유형", "TASK", "요청일", "요청자", "결과"]
                missing_columns = [col for col in required_columns if col not in df.columns]
                
                if missing_columns:
                    st.error(f"업로드한 엑셀 파일에 다음 필수 열이 없습니다: {', '.join(missing_columns)}")
                else:
                    # 업로드 버튼
                    if st.button("데이터 추가하기", key="sm_activity_upload_btn"):
                        # 현재 워크시트의 모든 데이터 가져오기
                        sheet_data = get_worksheet_data(worksheet)
                        # 헤더 행을 제외한 데이터 행 수 계산
                        current_row_count = len(sheet_data) - 1 if len(sheet_data) > 0 else 0
                        
                        # 진행 상황 표시
                        progress_bar = st.progress(0)
                        status_text = st.empty()
                        status_text.text("데이터 처리 중...")
                        
                        # 배치로 추가할 모든 행 준비
                        all_rows_to_add = []
                        error_rows = []
                        
                        # 각 행을 순회하면서 데이터 준비
                        for index, row in df.iterrows():
                            try:
                                # 진행 상황 업데이트 (10개 단위로 표시 업데이트)
                                if index % 10 == 0 or index == len(df) - 1:
                                    progress = (index + 1) / len(df)
                                    progress_bar.progress(progress)
                                    status_text.text(f"처리 중... {index + 1}/{len(df)}")
                                
                                # 요청일 처리 (날짜 형식 확인)
                                try:
                                    if pd.isna(row.get('요청일')):
                                        req_date = datetime.today()
                                    elif isinstance(row['요청일'], datetime):
                                        req_date = row['요청일']
                                    else:
                                        # 문자열인 경우 파싱 시도
                                        req_date = datetime.strptime(str(row['요청일']), "%Y-%m-%d")
                                except:
                                    req_date = datetime.today()
                                
                                # 작업일은 요청일과 동일하게 설정
                                work_date = req_date
                                
                                # 새 행 번호 계산
                                new_row_num = current_row_count + len(all_rows_to_add) + 1
                                
                                # 데이터 준비
                                new_row_data = [
                                    str(new_row_num),  # NO
                                    req_date.strftime("%Y%m"),  # 월 정보
                                    str(row.get('구분', '')),  # 구분
                                    str(row.get('작업유형', '')),  # 작업유형
                                    str(row.get('TASK', '')),  # TASK
                                    req_date.strftime("%Y-%m-%d"),  # 요청일
                                    work_date.strftime("%Y-%m-%d"),  # 작업일
                                    str(row.get('요청자', '')),  # 요청자
                                    str(row.get('IT', 'IT 담당자')),  # IT 담당자
                                    str(row.get('CNS', 'CNS 담당자')),  # CNS 담당자
                                    str(row.get('개발자', '개발자')),  # 개발자
                                    str(row.get('내용', row.get('TASK', ''))),  # 내용
                                    str(row.get('결과', '완료'))  # 결과
                                ]
                                
                                # 배열에 추가
                                all_rows_to_add.append(new_row_data)
                                
                            except Exception as e:
                                error_rows.append(index)
                                st.error(f"행 {index+1} 처리 중 오류 발생: {str(e)[:100]}...")
                        
                        # 배치 처리를 위한 상태 업데이트
                        status_text.text("Google 스프레드시트에 데이터 추가 중...")
                        
                        try:
                            # 배치 단위로 나누어 추가 (API 할당량 고려)
                            batch_size = 25  # 한 번에 추가할 최대 행 수 감소
                            success_count = 0
                            
                            for i in range(0, len(all_rows_to_add), batch_size):
                                batch = all_rows_to_add[i:i+batch_size]
                                if batch:
                                    # 배치 단위로 데이터 추가
                                    worksheet.append_rows(batch)
                                    success_count += len(batch)
                                    
                                    # 배치 추가 후 진행 상황 업데이트
                                    batch_progress = min(1.0, (i + len(batch)) / len(all_rows_to_add))
                                    progress_bar.progress(batch_progress)
                                    status_text.text(f"추가 중... {i + len(batch)}/{len(all_rows_to_add)} 행")
                                    
                                    # API 할당량 제한을 고려한 딜레이 (필요시)
                                    if i + batch_size < len(all_rows_to_add):
                                        status_text.text(f"API 할당량 제한 방지를 위해 잠시 대기 중... ({(i + len(batch))}/{len(all_rows_to_add)} 완료)")
                                        time.sleep(3)  # 3초로 대기 시간 증가
                            
                            # 진행 상황 완료
                            progress_bar.progress(1.0)
                            status_text.text("처리 완료! 데이터 정렬 중...")
                            
                            # 요청일 기준으로 데이터 정렬
                            try:
                                sort_worksheet_by_date(worksheet)
                                # 캐시 갱신 함수 호출
                                refresh_worksheet_data()
                                st.success(f"✅ 업로드 완료! 총 {success_count}개 행이 성공적으로 추가되었습니다. (오류: {len(error_rows)}개)")
                                if error_rows:
                                    st.warning(f"일부 행({len(error_rows)}개)에서 오류가 발생했습니다. 해당 행: {', '.join(map(str, [r+1 for r in error_rows]))}")
                                # 데이터 업데이트 후 자동 새로고침
                                st.rerun()
                            except Exception as e:
                                st.warning(f"데이터는 추가되었으나 정렬 중 오류가 발생했습니다: {str(e)[:150]}...")
                                st.info("API 할당량 제한으로 인한 오류일 수 있습니다. 1-2시간 후에 다시 시도하거나, 단일 항목을 추가하여 자동 정렬을 트리거할 수 있습니다.")
                            
                        except Exception as e:
                            st.error(f"데이터 배치 추가 중 오류가 발생했습니다: {str(e)[:200]}...")
                            st.info("Google Sheets API 할당량 제한으로 인한 오류일 수 있습니다. 다음 조치를 취하세요:")
                            st.markdown("""
                            1. 1-2시간 기다린 후 다시 시도하세요 (API 할당량이 재설정됨).
                            2. 더 작은 파일로 나누어 업로드하세요 (행 수를 줄임).
                            3. 단일 항목을 한 번에 하나씩 추가하세요.
                            """)
                            # 성공한 행 수가 있다면 표시
                            if success_count > 0:
                                st.info(f"{success_count}개 행은 성공적으로 추가되었습니다.")
            
            except Exception as e:
                st.error(f"파일 처리 중 오류가 발생했습니다: {str(e)}")

    # 폼 외부에 날짜 선택 UI 배치 (콜백 함수 사용 가능)
    st.subheader("📅 날짜 설정")

    date_col1, date_col2 = st.columns(2)
    with date_col1:
        # 요청일 선택 도움말 추가
        today = datetime.today()
        st.date_input(
            "요청일 선택", 
            key="req_date", 
            on_change=update_work_date,
            help="요청일을 선택하면 작업일이 자동으로 같은 날짜로 설정됩니다.",
            label_visibility="visible"
        )
    with date_col2:
        # 작업일 확인 도움말 추가
        st.date_input(
            "작업일 확인", 
            key="work_date", 
            disabled=True,
            help="요청일과 자동으로 동기화됩니다. 별도 변경은 불가능합니다.",
            label_visibility="visible"
        )

    # SM Activity 입력 양식 생성
    with st.form("activity_form"):
        # 각 필드 입력 UI 요소 생성
        st.subheader("📝 작업 정보 입력")
        
        구분 = st.selectbox("구분", ["정기", "비정기"])  # 작업 구분 선택
        # 작업 유형 선택 드롭다운
        작업유형 = st.selectbox("작업유형", [
            "조간점검", "재적재", "인프라 작업", "SI 지원", "ERRC",
            "CCB", "적재", "시스템 운영", "월정기작업", "인수인계"
        ])
        task = st.text_input("TASK 제목")  # 작업 제목 입력
        
        # 담당자 정보를 한 줄에 4개 컬럼으로 배치
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            요청자 = st.text_input("요청자")  # 요청자 입력
        with col2:
            it = st.text_input("IT 담당자", value="한상욱")  # IT 담당자 입력(기본값 설정)
        with col3:
            cns = st.text_input("CNS 담당자", value="이정인")  # CNS 담당자 입력(기본값 설정)
        with col4:
            개발자 = st.text_input("개발자", value="위승빈")  # 개발자 입력(기본값 설정)
        
        결과 = st.selectbox("결과", ["진행 중", "완료", "보류", "기타"])  # 작업 결과 상태 선택

        # 양식 제출 버튼 생성
        submitted = st.form_submit_button("추가하기")

        # 양식이 제출되면 실행되는 로직
        if submitted:
            try:
                # 입력값 검증
                if not task:
                    st.error("TASK 제목을 입력해주세요.")
                    st.stop()
                
                요청일 = st.session_state.req_date  # 폼 외부에서 설정한 요청일 사용
                작업일 = st.session_state.work_date  # 폼 외부에서 설정한 작업일 사용
                
                # 캐싱된 함수를 사용하여 데이터 가져오기
                sheet_data = get_worksheet_data(worksheet)
                # 헤더 행을 제외한 데이터 행 수 계산
                current_row_count = len(sheet_data) - 1 if len(sheet_data) > 0 else 0
                
                # 새 행 번호 계산
                new_row_num = current_row_count + 1
                
                # 데이터 준비
                new_row_data = [
                    str(new_row_num),  # NO
                    요청일.strftime("%Y%m"),  # 월 정보 (YYYYMM 형식)
                    구분,  # 구분
                    작업유형,  # 작업유형
                    task,  # TASK
                    요청일.strftime("%Y-%m-%d"),  # 요청일
                    작업일.strftime("%Y-%m-%d"),  # 작업일
                    요청자,  # 요청자
                    it,  # IT 담당자
                    cns,  # CNS 담당자
                    개발자,  # 개발자
                    task,  # 내용 (TASK와 동일하게 설정)
                    결과  # 결과
                ]
                
                # Google 스프레드시트에 데이터 추가
                with st.spinner("데이터 추가 중..."):
                    worksheet.append_row(new_row_data)
                    # 캐시 무효화 (데이터가 변경되었으므로)
                    get_worksheet_data.clear()
                    
                    # 요청일 기준으로 데이터 정렬
                    try:
                        sort_worksheet_by_date(worksheet)
                        # 캐시 갱신 함수 호출
                        refresh_worksheet_data()
                    except Exception as e:
                        st.warning(f"데이터는 추가되었으나 정렬 중 오류가 발생했습니다: {str(e)[:150]}...")
                
                # 성공 메시지 표시
                st.success(f"✅ {selected_sheet_name} 문서에 성공적으로 추가되었습니다.\n\n**추가된 작업:** {task}")
                
                # 데이터 업데이트 후 자동 새로고침
                st.rerun()
            except Exception as e:
                st.error(f"데이터 추가 중 오류가 발생했습니다: {e}")

    # 현재 워크시트의 모든 데이터 가져와서 표시
    try:
        # 캐싱된 함수를 사용하여 데이터 가져오기
        sheet_data = get_worksheet_data(worksheet)
        if len(sheet_data) > 1:  # 헤더 행을 제외하고 데이터가 있는 경우
            st.subheader("📊 현재 기록된 데이터")
            df = pd.DataFrame(sheet_data[1:], columns=sheet_data[0])
            st.dataframe(df)
            
            # 엑셀 파일로 변환하여 다운로드 버튼 제공
            excel_buffer = BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name=worksheet_name)
                
                # 엑셀 서식 설정
                workbook = writer.book
                worksheet_excel = writer.sheets[worksheet_name]
                
                # 헤더 스타일 설정
                for col_num, value in enumerate(df.columns.values, 1):
                    cell = worksheet_excel.cell(row=1, column=col_num)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # 열 너비 설정
                worksheet_excel.column_dimensions['E'].width = 30  # TASK 컬럼
                worksheet_excel.column_dimensions['F'].width = 15  # 요청일 컬럼
                worksheet_excel.column_dimensions['G'].width = 15  # 작업일 컬럼
                worksheet_excel.column_dimensions['L'].width = 40  # 내용 컬럼
            
            excel_buffer.seek(0)
            
            st.download_button(
                label=f"📥 {selected_sheet_name} SM Activity 엑셀 다운로드",
                data=excel_buffer,
                file_name=f"{google_sheet_name}_SM_Activity.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.info("아직 기록된 데이터가 없습니다. 위 양식을 통해 새 활동을 추가해주세요.")
    except Exception as e:
        st.error(f"데이터 조회 중 오류가 발생했습니다: {e}")

    # 도움말 섹션 추가
    with st.expander("ℹ️ 도움말 및 사용 방법"):
        st.markdown("""
        ### 사용 방법
        
        #### SM Activity 탭
        1. 작성할 문서 유형을 선택합니다.
        2. 요청일을 선택하면 작업일이 자동으로 설정됩니다.
        3. 작업 정보를 입력하고 '추가하기' 버튼을 클릭합니다.
        4. 입력된 데이터는 자동으로 날짜순 정렬됩니다.
        5. **일괄 업로드**: '엑셀 파일 업로드' 섹션을 통해 여러 SM 활동을 한 번에 추가할 수 있습니다. 필요한 열 형식은 샘플 템플릿을 참고하세요.
        
        #### 현업문의 탭
        1. 요청일을 선택하면 답변일이 자동으로 설정됩니다.
        2. 문의 정보를 입력하고 '추가하기' 버튼을 클릭합니다.
        3. 입력된 데이터는 자동으로 요청일 기준으로 정렬됩니다.
        4. **일괄 업로드**: '엑셀 파일 업로드' 섹션을 통해 여러 현업문의를 한 번에 추가할 수 있습니다. 필요한 열 형식은 샘플 템플릿을 참고하세요.
        
        #### 데이터 다운로드
        1. 각 탭에서 해당 데이터만 다운로드:
           - SM Activity 탭에서는 "SM Activity 엑셀 다운로드" 버튼을 클릭하여 SM Activity 데이터만 다운로드할 수 있습니다.
           - 현업문의 탭에서는 "현업문의 엑셀 다운로드" 버튼을 클릭하여 현업문의 데이터만 다운로드할 수 있습니다.
        
        2. 모든 데이터 통합 다운로드:
           - 상단의 '데이터 다운로드' 섹션을 클릭합니다.
           - '전체 데이터 엑셀 파일 다운로드' 버튼을 클릭하면 SM Activity와 현업문의 데이터가 하나의 엑셀 파일(여러 시트)로 다운로드됩니다.
        
        ### 엑셀 파일 업로드
        엑셀 파일을 통해 여러 데이터를 한 번에 추가할 수 있습니다:
        
        #### SM Activity 데이터
        1. SM Activity 탭에서 샘플 템플릿을 다운로드하여 형식을 확인합니다.
        2. 업로드할 엑셀 파일은 다음 열들을 포함해야 합니다:
           - **구분**: 정기/비정기
           - **작업유형**: 조간점검, 재적재 등
           - **TASK**: 작업 제목
           - **요청일**: 날짜 형식 (YYYY-MM-DD)
           - **요청자**: 요청자 이름
           - **결과**: 진행 중, 완료, 보류, 기타
        
        #### 현업문의 데이터
        1. 현업문의 탭에서 샘플 템플릿을 다운로드하여 형식을 확인합니다.
        2. 업로드할 엑셀 파일은 다음 열들을 포함해야 합니다:
           - **문의방법**: Social Desk, MAIL, 메신저, 전화
           - **문의유형**: 개발사전검토, 데이터확인 등
           - **요청부서**: 부서명
           - **문의사항**: 문의 내용
           - **요청일**: 날짜 형식 (YYYY-MM-DD)
           - **요청자**: 요청자 이름
        
        ### 주의사항
        - 데이터는 Google 스프레드시트에 저장되며, 권한이 있는 사용자만 접근할 수 있습니다.
        - 대량의 데이터를 업로드할 경우 시간이 다소 소요될 수 있습니다.
        - 각 탭에서는 해당 탭에 맞는 데이터만 업로드해야 합니다. (SM Activity 탭에서는 SM Activity 데이터, 현업문의 탭에서는 현업문의 데이터)
        - 문제가 발생하면 관리자에게 문의하세요.
        """)

with tab2:
    # 현업문의 탭을 선택했음을 세션 상태에 저장
    update_current_tab("현업문의")
    
    # 현업문의 탭 내용
    st.subheader("📞 현업문의 기록")
    
    # 엑셀 파일 업로드 섹션 추가
    st.subheader("📤 엑셀 파일 업로드")
    with st.expander("엑셀 파일을 업로드하여 문의 데이터 일괄 추가"):
        # 샘플 템플릿 다운로드 기능 추가
        st.markdown("#### 샘플 템플릿 다운로드")
        inquiry_sample_df = pd.DataFrame({
            '문의방법': ['Social Desk', 'MAIL', '메신저', '전화'],
            '문의유형': ['개발사전검토', '데이터확인', '접속/권한문의', '공통'],
            '요청부서': ['인사팀', '마케팅팀', '영업팀', 'IT팀'],
            '문의사항': ['시스템 접근 권한 요청', '데이터 오류 확인', '기능 사용법 문의', '시스템 오류 보고'],
            '요청일': [datetime.today().strftime("%Y-%m-%d"), (datetime.today() - pd.Timedelta(days=1)).strftime("%Y-%m-%d"),
                     (datetime.today() - pd.Timedelta(days=2)).strftime("%Y-%m-%d"), (datetime.today() - pd.Timedelta(days=3)).strftime("%Y-%m-%d")],
            '요청자': ['홍길동', '김철수', '이영희', '박민수'],
            'IT': ['한상욱', '한상욱', '한상욱', '한상욱'],
            'CNS': ['이정인', '이정인', '이정인', '이정인'],
            '개발자': ['위승빈', '위승빈', '위승빈', '위승빈']
        })
        
        # 샘플 템플릿을 엑셀로 변환
        inquiry_sample_buffer = BytesIO()
        with pd.ExcelWriter(inquiry_sample_buffer, engine='openpyxl') as writer:
            inquiry_sample_df.to_excel(writer, index=False, sheet_name='현업문의')
        inquiry_sample_buffer.seek(0)
        
        # 샘플 템플릿 다운로드 버튼
        st.download_button(
            label="📝 샘플 템플릿 다운로드",
            data=inquiry_sample_buffer,
            file_name="현업문의_Template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="업로드 양식에 맞는 샘플 엑셀 템플릿을 다운로드합니다."
        )
        
        st.markdown("---")
        st.markdown("#### 데이터 업로드")
        inquiry_uploaded_file = st.file_uploader("현업문의 양식의 엑셀 파일을 업로드하세요", type=["xlsx", "xls"], key="inquiry_uploader")
        
        if inquiry_uploaded_file is not None:
            try:
                # 엑셀 파일 읽기
                inquiry_df = pd.read_excel(inquiry_uploaded_file, sheet_name=0)
                
                # 데이터프레임 미리보기 
                st.write("업로드한 데이터 미리보기:")
                st.dataframe(inquiry_df.head(5))
                
                # 필요한 열이 있는지 확인
                required_columns = ["문의방법", "문의유형", "요청부서", "문의사항", "요청일", "요청자"]
                missing_columns = [col for col in required_columns if col not in inquiry_df.columns]
                
                if missing_columns:
                    st.error(f"업로드한 엑셀 파일에 다음 필수 열이 없습니다: {', '.join(missing_columns)}")
                else:
                    # 업로드 버튼
                    if st.button("데이터 추가하기", key="inquiry_upload_btn"):
                        # 현재 워크시트의 모든 데이터 가져오기
                        inquiry_sheet_data = get_worksheet_data(inquiry_worksheet)
                        # 헤더 행을 제외한 데이터 행 수 계산
                        current_row_count = len(inquiry_sheet_data) - 1 if len(inquiry_sheet_data) > 0 else 0
                        
                        # 진행 상황 표시
                        progress_bar = st.progress(0)
                        status_text = st.empty()
                        status_text.text("데이터 처리 중...")
                        
                        # 배치로 추가할 모든 행 준비
                        all_rows_to_add = []
                        error_rows = []
                        
                        # 각 행을 순회하면서 데이터 준비
                        for index, row in inquiry_df.iterrows():
                            try:
                                # 진행 상황 업데이트 (10개 단위로 표시 업데이트)
                                if index % 10 == 0 or index == len(inquiry_df) - 1:
                                    progress = (index + 1) / len(inquiry_df)
                                    progress_bar.progress(progress)
                                    status_text.text(f"처리 중... {index + 1}/{len(inquiry_df)}")
                                
                                # 요청일 처리 (날짜 형식 확인)
                                try:
                                    if pd.isna(row.get('요청일')):
                                        req_date = datetime.today()
                                    elif isinstance(row['요청일'], datetime):
                                        req_date = row['요청일']
                                    else:
                                        # 문자열인 경우 파싱 시도
                                        req_date = datetime.strptime(str(row['요청일']), "%Y-%m-%d")
                                except:
                                    req_date = datetime.today()
                                
                                # 답변일은 요청일과 동일하게 설정
                                resp_date = req_date
                                
                                # 새 행 번호 계산
                                new_row_num = current_row_count + len(all_rows_to_add) + 1
                                
                                # 데이터 준비
                                new_row_data = [
                                    str(new_row_num),  # NO
                                    req_date.strftime("%Y%m"),  # 월 정보
                                    str(row.get('문의방법', 'Social Desk')),  # 문의방법
                                    str(row.get('문의유형', '데이터확인')),  # 문의유형
                                    str(row.get('요청부서', '')),  # 요청부서
                                    str(row.get('문의사항', '')),  # 문의사항
                                    req_date.strftime("%Y-%m-%d"),  # 요청일
                                    resp_date.strftime("%Y-%m-%d"),  # 답변일
                                    str(row.get('요청자', '')),  # 요청자
                                    str(row.get('IT', '한상욱')),  # IT 담당자
                                    str(row.get('CNS', '이정인')),  # CNS 담당자
                                    str(row.get('개발자', '위승빈'))  # 개발자
                                ]
                                
                                # 배열에 추가
                                all_rows_to_add.append(new_row_data)
                                
                            except Exception as e:
                                error_rows.append(index)
                                st.error(f"행 {index+1} 처리 중 오류 발생: {str(e)[:100]}...")
                        
                        # 배치 처리를 위한 상태 업데이트
                        status_text.text("Google 스프레드시트에 데이터 추가 중...")
                        
                        try:
                            # 배치 단위로 나누어 추가 (API 할당량 고려)
                            batch_size = 25  # 한 번에 추가할 최대 행 수 감소
                            success_count = 0
                            
                            for i in range(0, len(all_rows_to_add), batch_size):
                                batch = all_rows_to_add[i:i+batch_size]
                                if batch:
                                    # 배치 단위로 데이터 추가
                                    inquiry_worksheet.append_rows(batch)
                                    success_count += len(batch)
                                    
                                    # 배치 추가 후 진행 상황 업데이트
                                    batch_progress = min(1.0, (i + len(batch)) / len(all_rows_to_add))
                                    progress_bar.progress(batch_progress)
                                    status_text.text(f"추가 중... {i + len(batch)}/{len(all_rows_to_add)} 행")
                                    
                                    # API 할당량 제한을 고려한 딜레이 (필요시)
                                    if i + batch_size < len(all_rows_to_add):
                                        status_text.text(f"API 할당량 제한 방지를 위해 잠시 대기 중... ({(i + len(batch))}/{len(all_rows_to_add)} 완료)")
                                        time.sleep(3)  # 3초로 대기 시간 증가
                            
                            # 진행 상황 완료
                            progress_bar.progress(1.0)
                            status_text.text("처리 완료! 데이터 정렬 중...")
                            
                            # 요청일 기준으로 데이터 정렬
                            try:
                                sort_worksheet_by_date(inquiry_worksheet, date_col_idx=6)  # 요청일 열 인덱스가 6번째
                                # 캐시 갱신 함수 호출
                                refresh_worksheet_data()
                                st.success(f"✅ 업로드 완료! 총 {success_count}개 행이 성공적으로 추가되었습니다. (오류: {len(error_rows)}개)")
                                if error_rows:
                                    st.warning(f"일부 행({len(error_rows)}개)에서 오류가 발생했습니다. 해당 행: {', '.join(map(str, [r+1 for r in error_rows]))}")
                                # 데이터 업데이트 후 자동 새로고침
                                st.rerun()
                            except Exception as e:
                                st.warning(f"데이터는 추가되었으나 정렬 중 오류가 발생했습니다: {str(e)[:150]}...")
                                st.info("API 할당량 제한으로 인한 오류일 수 있습니다. 1-2시간 후에 다시 시도하거나, 단일 항목을 추가하여 자동 정렬을 트리거할 수 있습니다.")
                            
                        except Exception as e:
                            st.error(f"데이터 배치 추가 중 오류가 발생했습니다: {str(e)[:200]}...")
                            st.info("Google Sheets API 할당량 제한으로 인한 오류일 수 있습니다. 다음 조치를 취하세요:")
                            st.markdown("""
                            1. 1-2시간 기다린 후 다시 시도하세요 (API 할당량이 재설정됨).
                            2. 더 작은 파일로 나누어 업로드하세요 (행 수를 줄임).
                            3. 단일 항목을 한 번에 하나씩 추가하세요.
                            """)
                            # 성공한 행 수가 있다면 표시
                            if success_count > 0:
                                st.info(f"{success_count}개 행은 성공적으로 추가되었습니다.")
            
            except Exception as e:
                st.error(f"파일 처리 중 오류가 발생했습니다: {str(e)}")
                
    # 폼 외부에 날짜 선택 UI 배치
    st.subheader("📅 날짜 설정")

    date_col1, date_col2 = st.columns(2)
    # 세션 상태 초기화 - 요청일과 답변일을 위한 설정은 이미 위에서 수행했으므로 제거

    # 요청일이 변경될 때 답변일을 업데이트하는 콜백 함수는 이미 위에서 정의했으므로 제거

    date_col1, date_col2 = st.columns(2)
    with date_col1:
        st.date_input(
            "요청일 선택", 
            key="inquiry_req_date", 
            on_change=update_inquiry_resp_date,
            help="요청일을 선택하면 답변일이 자동으로 같은 날짜로 설정됩니다.",
            label_visibility="visible"
        )
    with date_col2:
        st.date_input(
            "답변일 확인", 
            key="inquiry_resp_date", 
            disabled=True,
            help="요청일과 자동으로 동기화됩니다. 별도 변경은 불가능합니다.",
            label_visibility="visible"
        )

    # 현업문의 입력 양식 생성
    with st.form("inquiry_form"):
        st.subheader("📝 문의 정보 입력")
        
        문의방법 = st.selectbox("문의방법", ["Social Desk", "MAIL", "메신저", "전화"])
        문의유형 = st.selectbox("문의유형", [
            "개발사전검토", "데이터확인", "접속/권한문의", "공통", 
            "데이터셋업", "데이터재적재", "기능문의"
        ])
        요청부서 = st.text_input("요청부서")
        문의사항 = st.text_area("문의사항")
        
        # 담당자 정보를 한 줄에 4개 컬럼으로 배치
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            요청자 = st.text_input("요청자", key="inquiry_requestor")
        with col2:
            it = st.text_input("IT 담당자", value="한상욱", key="inquiry_it")
        with col3:
            cns = st.text_input("CNS 담당자", value="이정인", key="inquiry_cns")
        with col4:
            개발자 = st.text_input("개발자", value="위승빈", key="inquiry_dev")

        # 양식 제출 버튼 생성
        inquiry_submitted = st.form_submit_button("추가하기")

        # 양식이 제출되면 실행되는 로직
        if inquiry_submitted:
            try:
                # 입력값 검증
                if not 문의사항:
                    st.error("문의사항을 입력해주세요.")
                    st.stop()
                
                요청일 = st.session_state.inquiry_req_date
                답변일 = st.session_state.inquiry_resp_date
                
                # 캐싱된 함수를 사용하여 데이터 가져오기
                inquiry_sheet_data = get_worksheet_data(inquiry_worksheet)
                # 헤더 행을 제외한 데이터 행 수 계산
                current_row_count = len(inquiry_sheet_data) - 1 if len(inquiry_sheet_data) > 0 else 0
                
                # 새 행 번호 계산
                new_row_num = current_row_count + 1
                
                # 데이터 준비
                new_row_data = [
                    str(new_row_num),  # NO
                    요청일.strftime("%Y%m"),  # 월 정보 (YYYYMM 형식)
                    문의방법,  # 문의방법
                    문의유형,  # 문의유형
                    요청부서,  # 요청부서
                    문의사항,  # 문의사항
                    요청일.strftime("%Y-%m-%d"),  # 요청일
                    답변일.strftime("%Y-%m-%d"),  # 답변일
                    요청자,  # 요청자
                    it,  # IT 담당자
                    cns,  # CNS 담당자
                    개발자  # 개발자
                ]
                
                # Google 스프레드시트에 데이터 추가
                with st.spinner("데이터 추가 중..."):
                    inquiry_worksheet.append_row(new_row_data)
                    # 캐시 무효화 (데이터가 변경되었으므로)
                    get_worksheet_data.clear()
                    
                    # 요청일 기준으로 데이터 정렬
                    try:
                        sort_worksheet_by_date(inquiry_worksheet, date_col_idx=6)  # 요청일 열 인덱스가 6번째
                        # 캐시 갱신 함수 호출
                        refresh_worksheet_data()
                    except Exception as e:
                        st.warning(f"데이터는 추가되었으나 정렬 중 오류가 발생했습니다: {str(e)[:150]}...")
                
                # 성공 메시지 표시
                st.success(f"✅ {selected_sheet_name} 문서의 현업문의 시트에 성공적으로 추가되었습니다.\n\n**추가된 문의:** {문의사항[:30]}...")
                
                # 데이터 업데이트 후 자동 새로고침
                st.rerun()
            except Exception as e:
                st.error(f"데이터 추가 중 오류가 발생했습니다: {e}")

    # 현재 워크시트의 모든 데이터 가져와서 표시
    try:
        # 캐싱된 함수를 사용하여 데이터 가져오기
        inquiry_sheet_data = get_worksheet_data(inquiry_worksheet)
        if len(inquiry_sheet_data) > 1:  # 헤더 행을 제외하고 데이터가 있는 경우
            st.subheader("📊 현재 기록된 문의 데이터")
            inquiry_df = pd.DataFrame(inquiry_sheet_data[1:], columns=inquiry_sheet_data[0])
            st.dataframe(inquiry_df)
            
            # 엑셀 파일로 변환하여 다운로드 버튼 제공
            excel_buffer = BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                inquiry_df.to_excel(writer, index=False, sheet_name=inquiry_worksheet_name)
                
                # 엑셀 서식 설정
                workbook = writer.book
                worksheet_excel = writer.sheets[inquiry_worksheet_name]
                
                # 헤더 스타일 설정
                for col_num, value in enumerate(inquiry_df.columns.values, 1):
                    cell = worksheet_excel.cell(row=1, column=col_num)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # 열 너비 설정
                worksheet_excel.column_dimensions['E'].width = 20  # 요청부서 컬럼
                worksheet_excel.column_dimensions['F'].width = 40  # 문의사항 컬럼
                worksheet_excel.column_dimensions['G'].width = 15  # 요청일 컬럼
                worksheet_excel.column_dimensions['H'].width = 15  # 답변일 컬럼
            
            excel_buffer.seek(0)
            
            st.download_button(
                label=f"📥 {selected_sheet_name} 현업문의 엑셀 다운로드",
                data=excel_buffer,
                file_name=f"{google_sheet_name}_현업문의.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.info("아직 기록된 문의 데이터가 없습니다. 위 양식을 통해 새 문의를 추가해주세요.")
    except Exception as e:
        st.error(f"문의 데이터 조회 중 오류가 발생했습니다: {e}")

    # 도움말 섹션 추가
    with st.expander("ℹ️ 도움말 및 사용 방법"):
        st.markdown("""
        ### 사용 방법
        
        #### SM Activity 탭
        1. 작성할 문서 유형을 선택합니다.
        2. 요청일을 선택하면 작업일이 자동으로 설정됩니다.
        3. 작업 정보를 입력하고 '추가하기' 버튼을 클릭합니다.
        4. 입력된 데이터는 자동으로 날짜순 정렬됩니다.
        5. **일괄 업로드**: '엑셀 파일 업로드' 섹션을 통해 여러 SM 활동을 한 번에 추가할 수 있습니다. 필요한 열 형식은 샘플 템플릿을 참고하세요.
        
        #### 현업문의 탭
        1. 요청일을 선택하면 답변일이 자동으로 설정됩니다.
        2. 문의 정보를 입력하고 '추가하기' 버튼을 클릭합니다.
        3. 입력된 데이터는 자동으로 요청일 기준으로 정렬됩니다.
        4. **일괄 업로드**: '엑셀 파일 업로드' 섹션을 통해 여러 현업문의를 한 번에 추가할 수 있습니다. 필요한 열 형식은 샘플 템플릿을 참고하세요.
        
        #### 데이터 다운로드
        1. 각 탭에서 해당 데이터만 다운로드:
           - SM Activity 탭에서는 "SM Activity 엑셀 다운로드" 버튼을 클릭하여 SM Activity 데이터만 다운로드할 수 있습니다.
           - 현업문의 탭에서는 "현업문의 엑셀 다운로드" 버튼을 클릭하여 현업문의 데이터만 다운로드할 수 있습니다.
        
        2. 모든 데이터 통합 다운로드:
           - 상단의 '데이터 다운로드' 섹션을 클릭합니다.
           - '전체 데이터 엑셀 파일 다운로드' 버튼을 클릭하면 SM Activity와 현업문의 데이터가 하나의 엑셀 파일(여러 시트)로 다운로드됩니다.
        
        ### 엑셀 파일 업로드
        엑셀 파일을 통해 여러 데이터를 한 번에 추가할 수 있습니다:
        
        #### SM Activity 데이터
        1. SM Activity 탭에서 샘플 템플릿을 다운로드하여 형식을 확인합니다.
        2. 업로드할 엑셀 파일은 다음 열들을 포함해야 합니다:
           - **구분**: 정기/비정기
           - **작업유형**: 조간점검, 재적재 등
           - **TASK**: 작업 제목
           - **요청일**: 날짜 형식 (YYYY-MM-DD)
           - **요청자**: 요청자 이름
           - **결과**: 진행 중, 완료, 보류, 기타
        
        #### 현업문의 데이터
        1. 현업문의 탭에서 샘플 템플릿을 다운로드하여 형식을 확인합니다.
        2. 업로드할 엑셀 파일은 다음 열들을 포함해야 합니다:
           - **문의방법**: Social Desk, MAIL, 메신저, 전화
           - **문의유형**: 개발사전검토, 데이터확인 등
           - **요청부서**: 부서명
           - **문의사항**: 문의 내용
           - **요청일**: 날짜 형식 (YYYY-MM-DD)
           - **요청자**: 요청자 이름
        
        ### 주의사항
        - 데이터는 Google 스프레드시트에 저장되며, 권한이 있는 사용자만 접근할 수 있습니다.
        - 대량의 데이터를 업로드할 경우 시간이 다소 소요될 수 있습니다.
        - 각 탭에서는 해당 탭에 맞는 데이터만 업로드해야 합니다. (SM Activity 탭에서는 SM Activity 데이터, 현업문의 탭에서는 현업문의 데이터)
        - 문제가 발생하면 관리자에게 문의하세요.
        """)
